import io
import os
import shutil
import zipfile
import subprocess
import tempfile
import traceback
import openpyxl
from openpyxl.utils import get_column_letter
import pandas as pd
import matplotlib.pyplot as plt
import matplotlib
import img2pdf
from PIL import Image
import streamlit as st

# ==================== 全局配置 ====================
A4_WIDTH_PT  = 595.276
A4_HEIGHT_PT = 841.890

# 设置 matplotlib 中文字体
def set_chinese_font():
    try:
        font_path = '/usr/share/fonts/opentype/noto/NotoSansCJK-Regular.ttc'
        if os.path.exists(font_path):
            from matplotlib.font_manager import FontProperties
            font_prop = FontProperties(fname=font_path)
            matplotlib.rcParams['font.family'] = font_prop.get_name()
        else:
            matplotlib.rcParams['font.sans-serif'] = ['Noto Sans CJK JP', 'Noto Sans CJK SC', 'WenQuanYi Zen Hei', 'SimHei', 'DejaVu Sans']
            matplotlib.rcParams['axes.unicode_minus'] = False
    except:
        pass

set_chinese_font()

# ==================== LibreOffice 检测与转换 ====================
def get_soffice_path():
    paths = [
        '/usr/bin/soffice',
        '/Applications/LibreOffice.app/Contents/MacOS/soffice',
        'soffice'
    ]
    for p in paths:
        if os.path.exists(p) or shutil.which(p):
            return p
    return None

def try_libreoffice(file_bytes, base_name):
    """尝试用 LibreOffice 转换，失败返回 None"""
    soffice = get_soffice_path()
    if not soffice:
        return None
    try:
        with tempfile.TemporaryDirectory() as tmpdir:
            xlsx_path = os.path.join(tmpdir, f'{base_name}.xlsx')
            with open(xlsx_path, 'wb') as f:
                f.write(file_bytes)
            subprocess.run(
                [soffice, '--headless', '--convert-to', 'pdf', '--outdir', tmpdir, xlsx_path],
                check=True, stdout=subprocess.DEVNULL, stderr=subprocess.DEVNULL,
                timeout=30
            )
            pdf_path = os.path.join(tmpdir, f'{base_name}.pdf')
            if os.path.exists(pdf_path):
                with open(pdf_path, 'rb') as f:
                    return f.read()
    except Exception:
        pass
    return None

def get_sheet_names(file_bytes):
    """读取所有工作表名称"""
    wb = openpyxl.load_workbook(io.BytesIO(file_bytes), read_only=True, data_only=True)
    names = wb.sheetnames
    wb.close()
    return names

def split_pdf_pages(pdf_bytes, base_name, sheet_names):
    """拆分多页 PDF，用工作表名命名"""
    try:
        from PyPDF2 import PdfReader, PdfWriter
    except ImportError:
        return {f'{base_name}.pdf': pdf_bytes}
    reader = PdfReader(io.BytesIO(pdf_bytes))
    total = len(reader.pages)
    if total == 0:
        return {}
    pdfs = {}
    for i in range(total):
        writer = PdfWriter()
        writer.add_page(reader.pages[i])
        buf = io.BytesIO()
        writer.write(buf)
        buf.seek(0)
        name = f'{base_name}_{sheet_names[i]}.pdf' if i < len(sheet_names) else f'{base_name}_page{i+1}.pdf'
        pdfs[name] = buf.read()
    return pdfs

# ==================== Matplotlib 回退方案 ====================
def get_column_widths_from_excel(file_bytes):
    """读取 Excel 列宽（安全处理合并单元格）"""
    wb = openpyxl.load_workbook(io.BytesIO(file_bytes))
    ws = wb.active
    widths = []
    max_col = ws.max_column or 0
    for col_idx in range(1, max_col + 1):
        col_letter = get_column_letter(col_idx)
        col_dim = ws.column_dimensions.get(col_letter)
        w = col_dim.width if col_dim and col_dim.width else 8.43
        widths.append(w)
    wb.close()
    return widths

def sheet_to_pdf_matplotlib(df, col_widths=None):
    """将 DataFrame 转为 A4 单页 PDF（含中文支持）"""
    raw = df.values.tolist()
    data = [[str(cell) for cell in row] for row in raw]
    nrows = len(data)
    ncols = len(data[0]) if nrows > 0 else 0
    if col_widths is None or len(col_widths) != ncols:
        col_widths = [1.0] * ncols
    total_width = sum(col_widths)
    col_ratios = [w / total_width for w in col_widths]

    # 自适应字体大小
    margin = 20
    avail_width = A4_WIDTH_PT - margin
    max_chars = [max((len(data[r][c]) for r in range(nrows)), default=0) for c in range(ncols)]
    font_size_w = min([avail_width / (max_chars[i] * 0.5) / (col_ratios[i] * ncols) if max_chars[i] > 0 else 12 for i in range(ncols)])
    avail_height = A4_HEIGHT_PT - margin
    font_size_h = avail_height / nrows / 1.4
    font_size = min(font_size_w, font_size_h, 10)
    font_size = max(font_size, 3)

    fig, ax = plt.subplots(figsize=(8.27, 11.69))
    ax.axis('off')
    fig.subplots_adjust(left=0.03, right=0.97, bottom=0.03, top=0.97)
    table = ax.table(cellText=data, bbox=[0, 0, 1, 1])
    table.auto_set_font_size(False)
    table.set_fontsize(font_size)
    for (row, col), cell in table.get_celld().items():
        cell.set_width(col_ratios[col])
        cell.set_height(1.0 / nrows)
        cell.set_linewidth(0.3)
        cell.set_edgecolor('#CCCCCC')
        cell.set_text_props(ha='center', va='center')

    buf = io.BytesIO()
    try:
        fig.savefig(buf, format='png', dpi=200, pad_inches=0, bbox_inches='tight')
    finally:
        plt.close(fig)
    buf.seek(0)

    image = Image.open(buf)
    if image.mode in ('RGBA', 'LA', 'P'):
        rgb = Image.new('RGB', image.size, (255, 255, 255))
        rgb.paste(image, mask=image.split()[-1] if image.mode == 'RGBA' else None)
        image = rgb

    img_bytes = io.BytesIO()
    image.save(img_bytes, format='PNG')
    img_bytes.seek(0)

    layout = img2pdf.get_layout_fun(pagesize=(A4_WIDTH_PT, A4_HEIGHT_PT))
    return img2pdf.convert(img_bytes.getvalue(), layout_fun=layout)

# ==================== Streamlit 界面 ====================
st.set_page_config(page_title='Excel → A4 PDF')
st.title('📄 Excel 批量转 A4 单页 PDF')
st.markdown(
    '每个工作表生成一个 PDF，自动缩放。'
    '优先使用 LibreOffice 完美打印，否则自动回退到 matplotlib（已支持中文）。'
)

uploaded_files = st.file_uploader('选择 Excel 文件', type=['xlsx', 'xls'], accept_multiple_files=True)

if uploaded_files and st.button('🚀 开始转换'):
    with st.spinner('转换中…'):
        zip_buf = io.BytesIO()
        total_pdfs = 0
        libre_available = get_soffice_path() is not None
        if not libre_available:
            st.info('ℹ️ LibreOffice 未安装，使用 matplotlib 转换（带中文支持）。')

        with zipfile.ZipFile(zip_buf, 'w', zipfile.ZIP_DEFLATED) as zf:
            for f in uploaded_files:
                try:
                    raw = f.read()
                    name_no_ext = os.path.splitext(f.name)[0]
                    sheet_names = get_sheet_names(raw)
                    pdfs = {}

                    if libre_available:
                        pdf_bytes = try_libreoffice(raw, name_no_ext)
                        if pdf_bytes is not None:
                            pdfs = split_pdf_pages(pdf_bytes, name_no_ext, sheet_names)
                        else:
                            libre_available = False
                            st.warning('⚠️ LibreOffice 转换失败，回退到 matplotlib')

                    if not libre_available or not pdfs:
                        # matplotlib 回退
                        col_widths = get_column_widths_from_excel(raw)
                        xls = pd.ExcelFile(io.BytesIO(raw))
                        pdfs = {}
                        for sheet in sheet_names:
                            df = pd.read_excel(xls, sheet_name=sheet, header=None, dtype=str, keep_default_na=False)
                            if df.empty or df.size == 0:
                                continue
                            pdf_bytes = sheet_to_pdf_matplotlib(df, col_widths)
                            pdf_name = f'{name_no_ext}_{sheet}.pdf' if len(sheet_names) > 1 else f'{name_no_ext}.pdf'
                            pdfs[pdf_name] = pdf_bytes
                        xls.close()

                    for pname, pdata in pdfs.items():
                        zf.writestr(pname, pdata)
                        total_pdfs += 1
                except Exception as e:
                    st.error(f'❌ {f.name}\n{e}\n{traceback.format_exc()}')
        zip_buf.seek(0)
        if total_pdfs > 0:
            st.success(f'✅ 转换完成，共生成 {total_pdfs} 个 PDF')
            st.download_button('⬇️ 下载 ZIP', data=zip_buf, file_name='excel_pdfs.zip', mime='application/zip')
        else:
            st.error('❌ 未生成任何 PDF')
